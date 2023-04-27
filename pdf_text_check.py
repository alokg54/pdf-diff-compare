"""
Module to compare plain text of pdf files and write information to excel
output file.

"""
import pandas as pd
import difflib
import os
import os.path
from os import path
import time
import math
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Font, colors, PatternFill
from extract_text_util import pdftoText
import log_util
log = log_util.configure_logger()


class PdfCompare():
    """
    Class to compare plain text of new pdf and old pdf files and generate
    excel file to write extracted information

    """
    def __init__(self, file_path):
        """
        Init method of Pdfcompare class

        """
        self.pdf_text = pdftoText()
        self.input_excel_file_path = file_path

    def excel_load(self, file_name, column_to_load):
        """
        Method to read excel data using pandas

        :param column_to_load: `` list of column to read``
        :return: ``dataframe``
        """
        try:
            if os.path.exists(file_name):
                data_frame = pd.read_excel(file_name, usecols=column_to_load)
                # Add New required columns to data frame
                data_frame[['DELETED FROM OLD FILE', 'ADDED INTO NEW FILE',
                           'OLD FILE PAGE COUNT', 'NEW FILE PAGE COUNT',
                           'COMPARISON STATUS']] =\
                    pd.DataFrame([["", "", "", "", ""]])
                return data_frame
            else:
                raise FileNotFoundError(f'File doesn\'t exists: {file_name}')
        except FileNotFoundError as file_error:
            log.error(file_error)
            raise file_error
        except Exception as exp:
            log.error(exp)
            raise exp

    def compare_pdf(self, file1, file2, excel_out_file=None):
        """
        Method to compare two pdf file and shows a detailed difference in an
        output excel file

        :param excel_out_file: ``excel file data``
        :param file1: ``new pdf file``
        :param file2: ``old pdf file``
        :return: ``page count of file, data deleted from old file, data added into new file``
        """
        try:
            pdf_info = {}
            comparison_status = ""
            added_into_new_file, deleted_from_old_file = "", ""
            if not path.exists(file1) or not path.exists(file2):
                comparison_status = "FILE NOT FOUND"
            else:
                pdf1_text, new_file_page_count, pdf1_file_name = \
                    self.pdf_text.extract_text_pdf_py2pdf(file1)
                pdf2_text, old_file_page_count, pdf2_file_name = \
                    self.pdf_text.extract_text_pdf_py2pdf(file2)
                # check if file 1 or file 2 is read correctly
                if "Error" in pdf1_text or "Error" in pdf2_text:
                    pdf1_text, new_file_page_count, pdf1_file_name = \
                        self.pdf_text.extract_text_pdf_pdftohtml(file1)
                    pdf2_text, old_file_page_count, pdf2_file_name = \
                        self.pdf_text.extract_text_pdf_pdftohtml(file2)
                    comparison_status = pdf1_text
                    # set None value to pdf1 text
                    # pdf1_text = None
                if 'Error' in pdf1_text or 'Error' in pdf2_text:
                    comparison_status = "Unable to parse pdf"

                elif pdf1_text and pdf2_text:
                    added_into_new_file, deleted_from_old_file = \
                    self.get_file_difference(pdf1_text, pdf2_text)
                    if added_into_new_file or deleted_from_old_file:
                        comparison_status = "CHANGED"
                    else:
                        comparison_status = "NOT CHANGED"
                else:
                    comparison_status = "Unable to parse pdf"

                pdf_info.update({'OLD FILE PAGE COUNT': old_file_page_count})
                pdf_info.update({'NEW FILE PAGE COUNT': new_file_page_count})
            pdf_info.update({'COMPARISON STATUS': comparison_status})
            if not excel_out_file:
                return pdf_info, deleted_from_old_file, \
                       added_into_new_file
            # If output excel mentioned then write to the file and return none

            # self.write_in_excel(excel_out_file, pdf_info_1, pdf_info_2,
            #                     filtered_new_data, filtered_old_data)

        except Exception as exp:
            log.error(exp)
            # raise exp


    @staticmethod
    def get_pdf_info(pdf_filepath):
        """
        Method to get pdf information creation date, modification date and
        size of pdf file

        :param pdf_filepath: ``path of the pdf file``
        :return: ``modificationTime, creation_time, file_size_in_kb``
        """
        try:
            # Get Modified time
            mod_timesince_epoc = os.path.getmtime(pdf_filepath)
            modification_time = time.strftime("%Y-%m-%d  %H:%M:%S",
                                             time.localtime(mod_timesince_epoc))
            # Get creation time
            create_timesince_epoc = os.path.getctime(pdf_filepath)
            creation_time = time.strftime("%Y-%m-%d  %H:%M:%S",
                                          time.localtime(create_timesince_epoc))
            # get file size in kb
            file_size = os.path.getsize(pdf_filepath)
            # 1024 division added to get file size in kb
            file_size_in_kb = math.floor(file_size / 1024)
            file_info_dict = {'MODIFIED_DATE': modification_time,
                              'CREATION_DATE': creation_time, 'FILE_SIZE':
                                  str(file_size_in_kb) + " KB"}
            # file_info = "MODIFIED_DATE: " + modification_time + \
            #             "\n" + "CREATION_DATE: " + creation_time + "\n" \
            #             + "FILE_SIZE: " + str(file_size_in_kb) + "KB"
            return file_info_dict
        except Exception as exp:
            log.error("Error in retrieving file info ", exp)
            raise exp

    @staticmethod
    # TODO optimization required need to use pandas instead of openpyxl module

    def beatify_excel(excel_file):
        """
        Method to write data to excel file
        :param excel_file:
        :return: ``non``
        """
        try:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            # red_text = Font(color=colors.RED)
            # Yellow color background column
            yellow_background = PatternFill(bgColor=colors.YELLOW)
            diff_style_yellow = DifferentialStyle(fill=yellow_background)
            rule = Rule(type="expression", dxf=diff_style_yellow)
            rule.formula = ["$H1<3"]
            sheet.conditional_formatting.add("G1", rule)
            # Red color background column
            red_background = PatternFill(bgColor=colors.RED)
            diff_style_red = DifferentialStyle(fill=red_background)
            rule = Rule(type="expression", dxf=diff_style_red)
            rule.formula = ["$H1<3"]
            sheet.conditional_formatting.add("A1", rule)
            sheet.conditional_formatting.add("C1", rule)
            sheet.conditional_formatting.add("E1", rule)
            # Green color background column
            green_background = PatternFill(bgColor=colors.GREEN)
            diff_style_green = DifferentialStyle(fill=green_background)
            rule = Rule(type="expression", dxf=diff_style_green)
            rule.formula = ["$H1<3"]
            sheet.conditional_formatting.add("B1", rule)
            sheet.conditional_formatting.add("D1", rule)
            sheet.conditional_formatting.add("F1", rule)
            # sheet['C2'].font = Font(color='003311')
            # sheet['D2'].font = Font(color="cc0000")
            wb.save(excel_file)

        except Exception as exp:
            log.error("Exception occured: ", exp)
            raise exp

    def get_file_difference(self, data_pdf1, data_pdf2):
        """
        Method to calculate data added or removed in pdf1 compared to pdf2

        :param data_pdf1: ``data pdf1``
        :param data_pdf2: ``data pdf2``
        :return: ``data added in new file, data removed from old file``
        """
        try:
            data_pdf1 = data_pdf1.split(' ')
            data_pdf2 = data_pdf2.split(' ')
            data_added, data_removed = '', ''
            diff_raw = difflib.Differ()
            diff_array = list(diff_raw.compare(data_pdf1, data_pdf2))
            # Calculate data added
            data_added = "".join([word.replace('+', '') for word in
                                  diff_array if '+ ' in word])
            # for word in diff_array:
            #     if "+ " in word:
            #         data_added = data_added + word.replace("+", "")
            # check if data added is not none
            if data_added and data_added[-1] == ',':
                data_added = data_added.replace(',', '')
                data_added = data_added.strip()
            # Calculate data removed
            data_removed = "".join([word.replace('-', '') for word in
                                    diff_array if '- ' in word])
            # for word in diff_array:
            #     if "- " in word:
            #         data_removed = data_removed + word.replace("-", "")
            if data_removed and data_removed[-1] == ',':
                data_removed = data_removed.replace(',', '')
                data_removed = data_removed.strip()
            return data_added, data_removed
        except Exception as exp:
            log.error(exp)
            raise exp

    @staticmethod
    def _write_df_to_excel(df):
        """
        Method to write data frame into excel file

        :param df: ``data frame``
        :return: ``data frame``
        """
        try:
            for index, row in df.iterrows():
                file1, file2 = row['OLD FILE NAME'], row['NEW FILE NAME']
                if file1 and file2:
                    file1 = os.path.join(folder_path, file1)
                    file2 = os.path.join(folder_path, file2)
                pdf_info, deleted_from_old_file, added_to_new_file = \
                    pdfObj.compare_pdf(file1, file2)
                # any of the values is none Handle case in dataframe
                df.at[index, 'DELETED FROM OLD FILE'] = deleted_from_old_file
                df.at[index, 'ADDED INTO NEW FILE'] = added_to_new_file
                for key, value in pdf_info.items():
                    df.at[index, key] = value
            return df
        except Exception as exp:
            log.error(exp)
            # raise exp


if __name__ == '__main__':
    start = time.time()
    sheet_name = r'G:\PDF_Diff\TEST\Test_Data\01_Input_test.xlsx'
    output_file = r'G:\PDF_Diff\TEST\Test_Data\01_Output_test.xlsx'
    # sheet_name = r'G:\PDF_Diff\PDF_Files\01_Input_MURE_2000.xlsx'
    # output_file = r'G:\PDF_Diff\PDF_Files\01Output_New.xlsx'
    column_list = ['OLD FILE NAME', 'NEW FILE NAME']
    folder_path = os.path.dirname(sheet_name)
    pdfObj = PdfCompare(sheet_name)
    df = pdfObj.excel_load(sheet_name, column_list)
    data_frame = pdfObj._write_df_to_excel(df)
    data_frame.to_excel(output_file, index=False)
    pdfObj.beatify_excel(output_file)
    end = time.time()
    count_row = data_frame.shape[0]
    print("Total time to compare", count_row, "pair of files is:",
          "{:.2f}".format(end-start), "Sec.")
          